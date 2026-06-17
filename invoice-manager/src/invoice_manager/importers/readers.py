"""Read tabular invoice data from CSV and XLSX files."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path


@dataclass(slots=True)
class ImportedTable:
    headers: list[str]
    rows: list[dict[str, object]]


def read_import_file(content: bytes, filename: str) -> ImportedTable:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        return _read_csv(content)
    if suffix == ".xlsx":
        return _read_xlsx(content)
    raise ValueError("Obsługiwane formaty plików to CSV oraz XLSX.")


def _read_csv(content: bytes) -> ImportedTable:
    text: str | None = None
    for encoding in ("utf-8-sig", "cp1250"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Nie udało się odczytać kodowania pliku CSV.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("Plik CSV nie zawiera nagłówków.")
    source_headers = list(reader.fieldnames)
    headers = _clean_headers(source_headers)
    rows = [
        {
            header: source.get(source_header)
            for source_header, header in zip(source_headers, headers, strict=True)
        }
        for source in reader
        if any(str(value or "").strip() for value in source.values())
    ]
    return ImportedTable(headers=headers, rows=rows)


def _read_xlsx(content: bytes) -> ImportedTable:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "Obsługa XLSX wymaga pakietu openpyxl z requirements.txt."
        ) from error

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        values = worksheet.iter_rows(values_only=True)
        header_values = next(values, None)
        if header_values is None:
            raise ValueError("Arkusz XLSX jest pusty.")
        headers = _clean_headers(header_values)
        rows = [
            dict(zip(headers, row, strict=False))
            for row in values
            if any(value is not None and str(value).strip() for value in row)
        ]
        return ImportedTable(headers=headers, rows=rows)
    finally:
        workbook.close()


def _clean_headers(values: object) -> list[str]:
    headers: list[str] = []
    used: set[str] = set()
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value is not None else f"Kolumna {index}"
        header = base or f"Kolumna {index}"
        suffix = 2
        while header in used:
            header = f"{base} ({suffix})"
            suffix += 1
        used.add(header)
        headers.append(header)
    return headers
