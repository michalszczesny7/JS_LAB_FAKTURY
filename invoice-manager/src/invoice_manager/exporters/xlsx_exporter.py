"""Formatted multi-sheet XLSX export for invoice reports."""

from __future__ import annotations

from datetime import date
from io import BytesIO

from invoice_manager.exporters.csv_exporter import CSV_HEADERS
from invoice_manager.services.report_service import ReportData, ReportGroupRow

TITLE_FILL = "0F766E"
HEADER_FILL = "D1FAE5"
HEADER_TEXT = "134E4A"
MONEY_FORMAT = '#,##0.00 "PLN";[Red]-#,##0.00 "PLN"'
COUNT_FORMAT = "#,##0"


def export_report_xlsx(report: ReportData) -> bytes:
    """Return an XLSX workbook with detail, summary, and aggregate sheets."""

    try:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as error:
        raise RuntimeError(
            "Eksport XLSX wymaga pakietu openpyxl z requirements.txt."
        ) from error

    workbook = Workbook()
    invoices_sheet = workbook.active
    invoices_sheet.title = "Faktury"
    invoice_rows = [
        (
            row.id,
            row.invoice_number,
            date.fromisoformat(row.issue_date),
            date.fromisoformat(row.payment_date) if row.payment_date else None,
            row.contractor,
            row.investment,
            row.category,
            row.invoice_type,
            row.status,
            row.payment_status,
            row.net_amount,
            row.vat_amount,
            row.gross_amount,
            row.source_file,
        )
        for row in report.invoices
    ]
    _write_table(
        invoices_sheet,
        CSV_HEADERS,
        invoice_rows,
        "InvoicesTable",
        money_columns=(11, 12, 13),
        date_columns=(3, 4),
        Table=Table,
        TableStyleInfo=TableStyleInfo,
        PatternFill=PatternFill,
        Font=Font,
        Alignment=Alignment,
        Border=Border,
        Side=Side,
    )

    summary_sheet = workbook.create_sheet("Podsumowanie")
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.merge_cells("A1:B1")
    summary_sheet["A1"] = "Podsumowanie raportu faktur"
    summary_sheet["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    summary_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    summary_sheet["A1"].alignment = Alignment(horizontal="center")
    summary_values = (
        ("Liczba faktur", report.kpis.invoice_count, COUNT_FORMAT),
        ("Suma netto", report.kpis.net_total, MONEY_FORMAT),
        ("Suma VAT", report.kpis.vat_total, MONEY_FORMAT),
        ("Suma brutto", report.kpis.gross_total, MONEY_FORMAT),
        ("Zatwierdzone", report.kpis.approved_count, COUNT_FORMAT),
        ("Odrzucone", report.kpis.rejected_count, COUNT_FORMAT),
        ("Usunięte", report.kpis.deleted_count, COUNT_FORMAT),
    )
    for row_index, (label, value, number_format) in enumerate(
        summary_values, start=3
    ):
        summary_sheet.cell(row_index, 1, label)
        value_cell = summary_sheet.cell(row_index, 2, value)
        value_cell.number_format = number_format
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 20

    summary_sheet["D2"] = "Status"
    summary_sheet["E2"] = "Suma brutto"
    for cell in summary_sheet[2][3:5]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_TEXT, bold=True)
    for row_index, item in enumerate(report.by_status, start=3):
        summary_sheet.cell(row_index, 4, item.name)
        amount_cell = summary_sheet.cell(row_index, 5, item.gross_total)
        amount_cell.number_format = MONEY_FORMAT
    summary_sheet.column_dimensions["D"].width = 22
    summary_sheet.column_dimensions["E"].width = 20
    if report.by_status:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Suma brutto według statusu"
        chart.y_axis.title = "PLN"
        chart.x_axis.title = "Status"
        data = Reference(
            summary_sheet,
            min_col=5,
            min_row=2,
            max_row=2 + len(report.by_status),
        )
        categories = Reference(
            summary_sheet,
            min_col=4,
            min_row=3,
            max_row=2 + len(report.by_status),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 14
        summary_sheet.add_chart(chart, "G2")

    group_sheets = (
        ("Według inwestycji", "Inwestycja", report.by_investment, "InvestmentsTable"),
        ("Według kontrahentów", "Kontrahent", report.by_contractor, "ContractorsTable"),
        ("Miesięcznie", "Miesiąc", report.monthly, "MonthlyTable"),
        ("Według statusu", "Status", report.by_status, "StatusesTable"),
    )
    for sheet_name, label, items, table_name in group_sheets:
        worksheet = workbook.create_sheet(sheet_name)
        _write_group_sheet(
            worksheet,
            label,
            items,
            table_name,
            Table=Table,
            TableStyleInfo=TableStyleInfo,
            PatternFill=PatternFill,
            Font=Font,
            Alignment=Alignment,
            Border=Border,
            Side=Side,
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_group_sheet(
    worksheet: object,
    label: str,
    items: list[ReportGroupRow],
    table_name: str,
    **style_objects: object,
) -> None:
    headers = (label, "Liczba faktur", "Suma netto", "Suma VAT", "Suma brutto")
    rows = [
        (
            item.name,
            item.invoice_count,
            item.net_total,
            item.vat_total,
            item.gross_total,
        )
        for item in items
    ]
    _write_table(
        worksheet,
        headers,
        rows,
        table_name,
        money_columns=(3, 4, 5),
        count_columns=(2,),
        **style_objects,
    )


def _write_table(
    worksheet: object,
    headers: tuple[str, ...],
    rows: list[tuple[object, ...]],
    table_name: str,
    *,
    money_columns: tuple[int, ...] = (),
    count_columns: tuple[int, ...] = (),
    date_columns: tuple[int, ...] = (),
    **style_objects: object,
) -> None:
    Table = style_objects["Table"]
    TableStyleInfo = style_objects["TableStyleInfo"]
    PatternFill = style_objects["PatternFill"]
    Font = style_objects["Font"]
    Alignment = style_objects["Alignment"]
    Border = style_objects["Border"]
    Side = style_objects["Side"]

    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    thin_border = Border(bottom=Side(style="thin", color="D1D5DB"))
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_TEXT, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for column_index in money_columns:
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
        ):
            for item in cell:
                item.number_format = MONEY_FORMAT
    for column_index in count_columns:
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
        ):
            for item in cell:
                item.number_format = COUNT_FORMAT
    for column_index in date_columns:
        for cell in worksheet.iter_cols(
            min_col=column_index,
            max_col=column_index,
            min_row=2,
        ):
            for item in cell:
                item.number_format = "yyyy-mm-dd"

    for column_index, header in enumerate(headers, start=1):
        values = [str(header)] + [
            str(row[column_index - 1] or "") for row in rows
        ]
        width = min(max(max(len(value) for value in values) + 2, 12), 36)
        worksheet.column_dimensions[
            worksheet.cell(1, column_index).column_letter
        ].width = width

    if rows:
        reference = f"A1:{worksheet.cell(len(rows) + 1, len(headers)).coordinate}"
        table = Table(displayName=table_name, ref=reference)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
